import os
import openpyxl
import pandas as pd

from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from constants import RANKED_GAMES_SHEET, REGULAR_GAMES_SHEET

class ExcelDataSaver:
    def create_filename(self, player_id: str) -> str:
        """
        Generate a unique filename using player ID and current timestamp.
        """
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{player_id}_{current_time}.xlsx"
    
    def apply_formatting(self, workbook: openpyxl.Workbook) -> None:
        """
        Apply styling to all sheets in the workbook: header style, row colors, column width.
        """
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Header style setup
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_border = Border(*[Side(style="medium", color="000000")] * 4)

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows: alternating background colors and thin borders
            thin_border = Border(*[Side(style="thin", color="000000")] * 4)
            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                fill_color = "D9E1F2" if i % 2 == 0 else "FFFFFF"
                fill = PatternFill("solid", fgColor=fill_color)
                for cell in row:
                    cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-size column width based on content
            for col in sheet.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = (max_len + 2) * 1.2


    def export_to_excel(self, df: pd.DataFrame, df2: pd.DataFrame, player_id: str) -> str:
        """
        Export two DataFrames to a formatted Excel file.
        """
        filename = self.create_filename(player_id=player_id)

        try:
            # Save DataFrames to Excel without formatting
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=REGULAR_GAMES_SHEET, index=False)
                df2.to_excel(writer, sheet_name=RANKED_GAMES_SHEET, index=False)

            # Load the saved file and apply formatting
            workbook = openpyxl.load_workbook(filename)
            self.apply_formatting(workbook)

            # Save the formatted workbook again
            workbook.save(filename)

            file_path = os.path.abspath(filename)
            print(f"Excel file created: {file_path}")
            return file_path

        except Exception as e:
            print(f"Error while exporting Excel file: {e}")
            return ""